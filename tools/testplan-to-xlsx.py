"""Generate docs/manual_test_plan.xlsx from docs/manual_test_plan.md.

The markdown stays the source of truth — this is a projection of it into a form that is
actually pleasant to work through: one row per test, a Result dropdown instead of three
checkboxes, and a Summary sheet whose counts update live as you mark.

The dropdown is a real improvement over the markdown, not just a convenience: a cell holds
exactly one value, so the "more than one box ticked" case that testplan-tally.ps1 has to
detect and quarantine cannot occur at all.

Usage:  py -3 tools/testplan-to-xlsx.py [in.md] [out.xlsx]
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"

# Earth tones lifted from the app's own light theme so the workbook and SandiBumi look
# related rather than accidentally different.
INK = "332A1F"
DIM = "7C6B52"
RULE = "DDD0AF"
HEAD_BG = "E9DFC8"
INPUT_BG = "FFF7D6"  # the cells you fill
PASS_BG, FAIL_BG, BLOCK_BG = "D8EAD3", "F6D5D2", "FCE8C8"
PASS_FG, FAIL_FG, BLOCK_FG = "1E5E22", "9C2620", "8A5200"

# The coverage tag gets the strongest colour on YOURS, not on the covered states: the
# point of the column is to find the work nobody has done, and a sheet where the safe
# rows shout is a sheet you have to read past.
YOURS_BG, YOURS_FG = "FBE3C8", "8A4B12"
COVERED_FG = "7C6B52"

SECTION_RE = re.compile(r"^# Section (\S+)\s*[—-]\s*(.*)$")
TEST_RE = re.compile(r"^### (T-[A-Za-z]+-\d+)\s*[—-]\s*(.*)$")
# Heading tag added 2026-08-05: `### T-IMP-05 — ...  [YOURS]`. Split off the title so it
# becomes a filterable column rather than noise in the middle of a sentence.
TAG_RE = re.compile(r"\s*\[(YOURS|PART-AUTOMATED|GATE-PINNED)\]\s*$")


def clean(text: str) -> str:
    """Strip markdown bold markers; keep everything else readable as plain text."""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text, flags=re.S)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def field(block: str, label: str) -> str:
    """Pull `**Label:** value` up to the next bold label or end of block."""
    m = re.search(
        rf"\*\*{label}:?\*\*:?\s*(.*?)(?=\n\s*\*\*[A-Z]|\Z)", block, flags=re.S
    )
    return clean(m.group(1)) if m else ""


def parse(md_path: Path):
    lines = md_path.read_text(encoding="utf-8").splitlines()
    sections, tests = [], []
    sec_code = sec_name = ""
    cur = None
    buf: list[str] = []

    def flush():
        if cur is None:
            return
        block = "\n".join(buf)
        # Steps run from **Steps:** to the **Result — T-…:** anchor, and deliberately keep
        # the inline Expected/Known-issue lines: that is the order you read them in.
        m = re.search(r"\*\*Steps:\*\*\s*(.*?)(?=\*\*Result\s*[—-])", block, flags=re.S)
        cur["steps"] = clean(m.group(1)) if m else ""
        cur["tool"] = field(block, "Tool/panel")
        cur["pre"] = field(block, "Preconditions")
        cur["known"] = "Yes" if re.search(r"\*\*Known issue", block) else ""

        # Carry the marks across. A cell holds one value, so a test with MORE than one box
        # ticked is left BLANK rather than resolved — testplan-tally.ps1 quarantines that
        # case rather than scoring it, and picking one here would manufacture a result the
        # markdown never states.
        marks = re.findall(r"^- \[[xX]\]\s*(Pass|Fail|Blocked)", block, flags=re.M)
        cur["result"] = marks[0] if len(marks) == 1 else ""
        cur["contradictory"] = len(marks) > 1
        mn = re.search(r"^\*\*Notes:\*\*\s*(.*)$", block, flags=re.M)
        note = mn.group(1).strip() if mn else ""
        cur["note"] = "" if re.fullmatch(r"[_\\*\s]*", note) else note
        tests.append(cur)

    for line in lines:
        ms = SECTION_RE.match(line)
        if ms:
            flush()
            cur, buf = None, []
            sec_code, sec_name = ms.group(1), ms.group(2).strip()
            sections.append((sec_code, sec_name))
            continue
        mt = TEST_RE.match(line)
        if mt:
            flush()
            buf = []
            raw = mt.group(2).strip()
            tag = TAG_RE.search(raw)
            cur = {
                "sec": sec_code,
                "id": mt.group(1),
                "title": TAG_RE.sub("", raw).strip(),
                "cov": tag.group(1) if tag else "",
            }
            continue
        if cur is not None:
            buf.append(line)
    flush()
    return sections, tests


def style_header(ws, ncols):
    thin = Side(style="thin", color=RULE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color=INK)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 30


def prior_marks(out_path: Path):
    """Read Result/Notes already in the workbook, keyed by test id.

    Regenerating used to hand back an empty Result column, so any evening spent marking
    in Excel was destroyed by the next run of this script — silently, because the file it
    overwrote looked exactly like the one it wrote. The markdown cannot rescue that: it is
    the source of truth for the TESTS, but Excel-side marks were never written back to it.
    """
    if not out_path.exists():
        return {}
    try:
        from openpyxl import load_workbook
        ws = load_workbook(out_path, data_only=True)["Tests"]
    except Exception as exc:                      # a corrupt or hand-edited file
        print(f"note: could not read existing {out_path.name} ({exc}); starting clean")
        return {}
    head = [c.value for c in ws[1]]
    if "Test ID" not in head or "Result" not in head:
        return {}
    i_id, i_res = head.index("Test ID"), head.index("Result")
    i_note = head.index("Notes") if "Notes" in head else None
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[i_id]
        if not tid:
            continue
        res = (row[i_res] or "").strip() if isinstance(row[i_res], str) else ""
        note = ""
        if i_note is not None and isinstance(row[i_note], str):
            note = row[i_note].strip()
        if res or note:
            out[tid] = (res, note)
    return out


def build(md_path: Path, out_path: Path):
    sections, tests = parse(md_path)
    if len(tests) != 265:
        print(f"WARNING: parsed {len(tests)} tests, expected 265")

    # The markdown wins where it states a mark — it is the declared source of truth and
    # testplan-tally.ps1 scores it. Where it is silent, an existing workbook entry is
    # carried forward rather than wiped: silence in the markdown means "not recorded
    # there", never "not tested".
    prior = prior_marks(out_path)
    carried = 0
    for t in tests:
        old_res, old_note = prior.get(t["id"], ("", ""))
        if not t["result"] and old_res:
            t["result"] = old_res
            carried += 1
        if not t["note"] and old_note:
            t["note"] = old_note

    wb = Workbook()

    # ---------------------------------------------------------------- Tests
    ws = wb.active
    ws.title = "Tests"
    headers = [
        "Section", "Test ID", "Test", "Coverage", "Result", "Notes",
        "Known issue", "Tool / panel", "Preconditions", "Steps & expected",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    body = Font(name=FONT, size=10, color=INK)
    top = Alignment(vertical="top", wrap_text=True)
    for t in tests:
        ws.append([
            t["sec"], t["id"], t["title"], t["cov"], t["result"], t["note"],
            t["known"], t["tool"], t["pre"], t["steps"],
        ])

    last = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last, max_col=len(headers)):
        for cell in row:
            cell.font = body
            cell.alignment = top
        row[4].fill = PatternFill("solid", fgColor=INPUT_BG)   # Result
        row[5].fill = PatternFill("solid", fgColor=INPUT_BG)   # Notes
        row[3].alignment = Alignment(vertical="center", horizontal="center")
        row[4].alignment = Alignment(vertical="center", horizontal="center")
        row[6].alignment = Alignment(vertical="center", horizontal="center")

    widths = [9, 14, 40, 17, 11, 40, 9, 26, 34, 96]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cov_rng = f"D2:D{last}"
    ws.conditional_formatting.add(cov_rng, CellIsRule(
        operator="equal", formula=['"YOURS"'],
        fill=PatternFill("solid", fgColor=YOURS_BG),
        font=Font(name=FONT, bold=True, size=10, color=YOURS_FG)))
    for value in ('"PART-AUTOMATED"', '"GATE-PINNED"'):
        ws.conditional_formatting.add(cov_rng, CellIsRule(
            operator="equal", formula=[value],
            font=Font(name=FONT, size=10, color=COVERED_FG)))

    # showDropDown is inverted in the OOXML spec — the attribute means "suppress the in-cell
    # dropdown", so False is what puts the arrow in the cell. showErrorMessage must be set
    # explicitly: openpyxl leaves it off, and without it Excel accepts any typed value and the
    # list is merely advisory (verified via Excel COM: ShowError came back False).
    # showInputMessage stays off on purpose — a tooltip firing on every one of 265 cells during
    # a long click-through is an irritation, and the arrow is self-explanatory.
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
    )
    dv.errorStyle = "stop"
    dv.errorTitle = "Not a valid result"
    dv.error = "Pick Pass, Fail or Blocked from the dropdown, or leave the cell blank."
    ws.add_data_validation(dv)
    dv.add(f"E2:E{last}")

    rng = f"E2:E{last}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"Pass"'],
        fill=PatternFill("solid", fgColor=PASS_BG), font=Font(name=FONT, bold=True, color=PASS_FG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"Fail"'],
        fill=PatternFill("solid", fgColor=FAIL_BG), font=Font(name=FONT, bold=True, color=FAIL_FG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"Blocked"'],
        fill=PatternFill("solid", fgColor=BLOCK_BG), font=Font(name=FONT, bold=True, color=BLOCK_FG)))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = "E2"   # Section / ID / Test / Coverage stay put while you scroll right

    # -------------------------------------------------------------- Summary
    sm = wb.create_sheet("Summary")
    sm.append(["Code", "Section", "Tests", "Yours", "Yours left",
               "Pass", "Fail", "Blocked", "Untested", "Done"])
    style_header(sm, 10)

    n = len(tests)
    for i, (code, name) in enumerate(sections, start=2):
        sm.append([
            code, name,
            f'=COUNTIF(Tests!$A$2:$A${n + 1},$A{i})',
            f'=COUNTIFS(Tests!$A$2:$A${n + 1},$A{i},Tests!$D$2:$D${n + 1},"YOURS")',
            # "Yours left" counts a YOURS row whose Result is still blank — the honest
            # size of the queue, which the plain Yours count stops reflecting the moment
            # any of them is marked.
            f'=COUNTIFS(Tests!$A$2:$A${n + 1},$A{i},Tests!$D$2:$D${n + 1},"YOURS",'
            f'Tests!$E$2:$E${n + 1},"")',
            f'=COUNTIFS(Tests!$A$2:$A${n + 1},$A{i},Tests!$E$2:$E${n + 1},"Pass")',
            f'=COUNTIFS(Tests!$A$2:$A${n + 1},$A{i},Tests!$E$2:$E${n + 1},"Fail")',
            f'=COUNTIFS(Tests!$A$2:$A${n + 1},$A{i},Tests!$E$2:$E${n + 1},"Blocked")',
            f'=C{i}-F{i}-G{i}-H{i}',
            f'=IFERROR((F{i}+G{i}+H{i})/C{i},0)',
        ])
    tot = len(sections) + 2
    sm.append([
        "", "TOTAL",
        *[f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}{tot - 1})" for c in range(3, 10)],
        f'=IFERROR((F{tot}+G{tot}+H{tot})/C{tot},0)',
    ])

    for row in sm.iter_rows(min_row=2, max_row=tot, max_col=10):
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
        row[9].number_format = "0%"
    for c in range(1, 11):
        for cell in (sm.cell(row=tot, column=c),):
            cell.font = Font(name=FONT, size=10, bold=True, color=INK)
    for col, w in zip("ABCDEFGHIJ", [9, 46, 8, 8, 11, 8, 8, 10, 10, 8]):
        sm.column_dimensions[col].width = w
    sm.freeze_panes = "A2"

    sm.conditional_formatting.add(f"E2:E{tot}", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor=YOURS_BG),
        font=Font(name=FONT, bold=True, color=YOURS_FG)))
    sm.conditional_formatting.add(f"G2:G{tot}", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor=FAIL_BG), font=Font(name=FONT, bold=True, color=FAIL_FG)))
    sm.conditional_formatting.add(f"H2:H{tot}", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor=BLOCK_BG), font=Font(name=FONT, bold=True, color=BLOCK_FG)))

    # ------------------------------------------------------------ How to use
    hw = wb.create_sheet("How to use", 0)
    rows = [
        ("SandiBumi — manual test plan", ""),
        ("", ""),
        ("Generated from docs/manual_test_plan.md, which stays the source of truth.", ""),
        ("Regenerate any time with:  py -3 tools/testplan-to-xlsx.py", ""),
        ("", ""),
        ("WHICH CELLS YOU FILL IN", ""),
        ("The two cream-shaded columns on the Tests sheet, and nothing else:", ""),
        ("   Result (column E)", "pick Pass / Fail / Blocked from the dropdown"),
        ("   Notes  (column F)", "free text - what went wrong, or a timing for the PERF tests"),
        ("Everything else is reference text. Summary is all formulas - do not type in it.", ""),
        ("", ""),
        ("START WITH THE COVERAGE FILTER (COLUMN D)", ""),
        ("It says what automation already checks, and it is a priority order:", ""),
        ("   YOURS", "nothing automated touches it - start here"),
        ("   PART-AUTOMATED", "a test drives part of it; the markdown names the part that is not"),
        ("   GATE-PINNED", "the numbers run on every green gate; only the display is left"),
        ("Filter column D to YOURS and the sheet shows exactly the work nobody has done.", ""),
        ("A tag is never a pass. It says a machine checked something, not that the tool is", ""),
        ("the one you would reach for - only your tick says that.", ""),
        ("", ""),
        ("EXAMPLE OF A FILLED ROW", ""),
        ("Section", "SHELL"),
        ("Test ID", "T-SHELL-02"),
        ("Test", "Ribbon tab walk + overflow chevrons"),
        ("Coverage", "PART-AUTOMATED"),
        ("Result", "Fail"),
        ("Notes", "chevron did not appear until the window went below 900px"),
        ("", ""),
        ("HOW TO WORK THROUGH IT", ""),
        ("1. Filter Coverage to YOURS, then work one Section at a time.", ""),
        ("2. Mark exactly one result per test. Leave it blank until you have actually run it -", ""),
        ("   blank means untested, and the Summary counts it that way.", ""),
        ("3. A test whose precondition failed is Blocked, not Fail.", ""),
        ("4. Known issue = Yes means it is EXPECTED to fail in one specific way, already", ""),
        ("   confirmed in AUDIT-2026-07-21-full-qc.md. Mark it Fail and note 'known'. If it", ""),
        ("   fails a DIFFERENT way, that is a new finding - say so in Notes.", ""),
        ("5. Watch Summary as you go; it updates live.", ""),
        ("6. When done, filter Result to Fail and Blocked and send me that list.", ""),
        ("", ""),
        ("WHY THERE IS NO 'CONTRADICTORY' COLUMN", ""),
        ("The markdown version has three checkboxes per test, so it can be ticked", ""),
        ("inconsistently and the tally script has to quarantine those. A dropdown cell holds", ""),
        ("exactly one value, so that failure mode does not exist here.", ""),
    ]
    for a, b in rows:
        hw.append([a, b])
    hw["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
    for r in range(2, len(rows) + 1):
        a, b = hw.cell(row=r, column=1), hw.cell(row=r, column=2)
        # A heading is an ALL-CAPS line at the left margin. The indent test matters:
        # without it "   YOURS" in the coverage key reads as a section heading.
        text = a.value or ""
        heading = bool(text) and not text.startswith(" ") and text.isupper() and len(text) < 60
        a.font = Font(name=FONT, size=10, bold=heading, color=INK if heading else DIM)
        b.font = Font(name=FONT, size=10, color=INK)
    # Rows located by their own text rather than by index — this block has been renumbered
    # once already, and a hard-coded row number fails silently by shading the wrong line.
    labels = [r[0] for r in rows]
    for r, (a, b) in enumerate(rows, start=1):
        if a.startswith("   ") or (b and a in ("Section", "Test ID", "Test",
                                               "Coverage", "Result", "Notes")):
            hw.cell(row=r, column=1).font = Font(name=FONT, size=10, color=INK)
        if a in ("Result", "Notes") and b:
            hw.cell(row=r, column=2).fill = PatternFill("solid", fgColor=INPUT_BG)
    assert "Coverage" in labels, "the How-to-use example lost its Coverage row"
    hw.column_dimensions["A"].width = 78
    hw.column_dimensions["B"].width = 58
    hw.sheet_view.showGridLines = False

    wb.save(out_path)
    marked = sum(1 for t in tests if t["result"])
    clash = [t["id"] for t in tests if t["contradictory"]]
    print(f"wrote {out_path}  ({len(tests)} tests, {len(sections)} sections)")
    print(f"  coverage: " + ", ".join(
        f"{tag} {sum(1 for t in tests if t['cov'] == tag)}"
        for tag in ("YOURS", "PART-AUTOMATED", "GATE-PINNED")))
    print(f"  results carried in: {marked} ({carried} of them from the previous workbook)")
    if clash:
        print("  left blank - more than one box ticked in the markdown: " + ", ".join(clash))


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "docs" / "manual_test_plan.md"
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else root / "docs" / "manual_test_plan.xlsx"
    build(src, dst)
